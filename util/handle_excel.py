import openpyxl
import os
import sys

# 添加上层路径到搜索路径，即项目的根目录
sys.path.append('..')
from Base.baseCaseData import BaseCaseData


class HandleExcel:

    def get_sheet_data(self):
        '''
        获取sheet对象
        :return:表对象
        '''
        curPath = os.path.abspath(os.path.dirname(__file__))
        # 获取项目根路径，内容为当前项目的名字
        rootPath = curPath[:curPath.find('BaokangTest') + len('BaokangTest')]

        # 获取excel对象
        open_excel = openpyxl.load_workbook(rootPath + "/Case/BaokangTest.xlsx")
        sheetnames = open_excel.sheetnames
        # 按表名称获取表对象
        sheet_obj = open_excel[sheetnames[0]]
        return sheet_obj

    def get_cell_value(self, row, cols):
        '''
        获取一个单元格的内容
        :param row: 单元格第几行
        :param cols: 单元格第几列
        :return:单元格值
        '''
        cell_data = self.get_sheet_data().cell(row=row, column=cols).value
        return cell_data

    def get_rows_value(self, row):
        '''
        获取某一行的内容
        :param row:
        :return:
        '''
        return self.get_sheet_data()[row]

    def get_columns_value(self, key=None):
        '''
        获取某一列得数据
        '''
        columns_list = []
        if key == None:
            key = 'A'
        columns_list_data = self.get_sheet_data()[key]
        for i in columns_list_data:
            columns_list.append(i.value)
        return columns_list

    def get_rows_number(self, case_id):
        '''
        获取行号
        '''
        num = 1
        cols_data = self.get_columns_value(key='A')
        for col_data in cols_data:
            if case_id == col_data:
                return num
            num = num + 1
        return num

    def get_vale_by_number_and_field(self, caseNumber):
        '''
        通过用例编号获取数据对象
        :param caseNumber: 用例编号
        :return:数据对象
        '''
        num = self.get_rows_number(caseNumber)

        case_number = self.get_cell_value(num, 1)
        description = self.get_cell_value(num, 2)
        if_execute = self.get_cell_value(num, 3)
        precondition = self.get_cell_value(num, 4)
        depend_key = self.get_cell_value(num, 5)
        url = self.get_cell_value(num, 6)
        method = self.get_cell_value(num, 7)
        data = self.get_cell_value(num, 8)
        cookie = self.get_cell_value(num, 9)
        header = self.get_cell_value(num, 10)
        expect_method = self.get_cell_value(num, 11)
        expect_result = self.get_cell_value(num, 12)
        result = self.get_cell_value(num, 13)
        result_data = self.get_cell_value(num, 14)

        my_case_data = BaseCaseData(case_number, description, if_execute, precondition, depend_key, url, method, data,
                                    cookie, header, expect_method, expect_result, result, result_data)
        return my_case_data

# if __name__ == "__main__":
#     handle = HandleExcel()
#     res = handle.get_vale_by_number_and_field(1)
#     print(res)
